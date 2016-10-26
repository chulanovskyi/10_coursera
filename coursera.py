
import json
import random
from datetime import datetime

import requests
import babel
from lxml import etree as ET
from lxml import html as HT
from openpyxl import Workbook
from openpyxl.styles import Font


COURSE_COUNT = 20


def get_courses_urls():
    coursera_feed = 'https://www.coursera.org/sitemap~www~courses.xml'
    response = requests.get(coursera_feed)
    xml_root = ET.fromstring(response.content)
    xml_prefix = '{http://www.sitemaps.org/schemas/sitemap/0.9}'
    xml_url_nodes = xml_root.iterfind('.//%sloc' % xml_prefix)
    courses_urls = [clean_url.text for clean_url in xml_url_nodes]
    return courses_urls


def get_course_info(course_url):
    coursera_api = 'https://api.coursera.org/api/courses.v1'
    course_slug = course_url.split('/').pop()
    payload = {
        'q': 'slug',
        'slug': course_slug,
        'fields': 'workload,\
            primaryLanguages,\
            plannedLaunchDate,\
            upcomingSessionStartDate,\
            courseDerivatives.v1(averageFiveStarRating)',
        'includes': 'courseDerivatives'
    }
    api_response = requests.get(coursera_api, params=payload)
    api_json = json.loads(api_response.text)
    api_elements = api_json['elements'].pop()
    api_rating = api_json['linked']['courseDerivatives.v1'].pop()

    name = api_elements['name']
    language = get_languages(api_elements)
    start_date = get_start_date(api_elements)
    workload = get_workload(api_elements)
    rating = get_rating(api_rating)
    
    return [name, language, start_date, rating, workload]


def get_workload(json_data):
    workload = json_data['workload']
    if workload:
        return workload
    else:
        return 'Unknown'


def get_languages(json_data):
    lang_code = json_data['primaryLanguages'].pop()
    language = babel.Locale.parse(lang_code, sep='-').english_name
    return language


def get_start_date(json_data):
    try:
        launch_date = json_data['plannedLaunchDate']
    except KeyError:
        launch_date = None
    try:
        upcoming = json_data['upcomingSessionStartDate']/1000
        upcoming_normal = datetime.fromtimestamp(upcoming).strftime('%B %d %Y')
    except KeyError:
        upcoming_normal = None
    if upcoming_normal:
        return upcoming_normal
    else:
        return launch_date


def get_rating(json_data):
    try:
        return str(round(json_data['averageFiveStarRating'],1))
    except KeyError:
        return 'Not rated'


def output_courses_meta_to_xlsx(filepath, courses):
    column_names = ('Name', 'Language', 'Start date', 'Rating', 'Duration')
    col_size_lg, col_size_md, col_size_sm = (30, 20, 10)
    wb = Workbook()
    ws = wb.active
    ws.title = '%s random courses from Coursera' % COURSE_COUNT
    for cell_index, cell in enumerate(ws['A1:E1'][0]):
        cell.value = column_names[cell_index]
        cell.font = Font(bold=True)
    ws.column_dimensions['A'].width = col_size_lg
    ws.column_dimensions['B'].width = col_size_sm
    ws.column_dimensions['C'].width = col_size_md
    ws.column_dimensions['E'].width = col_size_lg
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=COURSE_COUNT+1):
        for cell_index, cell in enumerate(row):
            cell.value = courses[cell.row-2][cell_index]
    wb.save(filepath)


if __name__ == '__main__':
    print('Getting course list...')
    courses_urls = get_courses_urls()
    print('Done!')
    courses_meta = []
    print('Getting %d courses info...' % COURSE_COUNT)
    while len(courses_meta) != COURSE_COUNT:
        random_index = courses_urls.index(random.choice(courses_urls))
        random_course = courses_urls.pop(random_index)
        print('Get: %s' % random_course)
        course_info = get_course_info(random_course)
        courses_meta.append(course_info)
    print('Done!')
    print('Making excel file...')
    output_courses_meta_to_xlsx('Coursera.xlsx', courses_meta)
    print('Finish!')
